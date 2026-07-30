from __future__ import annotations

import base64
import hashlib
import json
import shutil
import subprocess
import sys
import tempfile
import unittest
import warnings
import zipfile
from pathlib import Path

from cryptography.hazmat.primitives import serialization
from cryptography.hazmat.primitives.asymmetric import ec
from openpyxl import Workbook


SCRIPT = Path(__file__).with_name("license_delivery.py")


class LicenseDeliveryCliTest(unittest.TestCase):
    def setUp(self) -> None:
        self.temp_dir = tempfile.TemporaryDirectory()
        self.root = Path(self.temp_dir.name)
        self.input_dir = self.root / "input"
        self.input_dir.mkdir()

    def tearDown(self) -> None:
        self.temp_dir.cleanup()

    def write_request(self, source_name: str, source_sha256: str) -> Path:
        request = {
            "schemaVersion": 1,
            "requestId": "REQ-DQ-COMMERCIAL-20260730-001",
            "licenseId": "LIC-DQ-COMMERCIAL-000001",
            "deliveryId": "DEL-DQ-COMMERCIAL-20260730-001",
            "customerId": "tdtech",
            "projectId": "dingqiao-commercial",
            "reason": "commercial-rollout",
            "issuedAt": "2026-07-30",
            "sources": [
                {
                    "fileName": source_name,
                    "sha256": source_sha256,
                    "sheets": [{"name": "Sheet1", "columns": ["SN"]}],
                }
            ],
            "policy": {
                "features": ["ASR", "TTS"],
                "sdkMajor": 1,
                "applicationRecord": {"mode": "none"},
                "certificateBinding": {"mode": "none"},
                "runtimeExpiry": {"mode": "perpetual"},
                "maintenance": {"mode": "unlimited"},
                "deviceBinding": "required",
            },
            "previousLicenseId": None,
        }
        path = self.root / "license-request.json"
        path.write_text(json.dumps(request), encoding="utf-8")
        return path

    def run_cli(self, *args: str) -> subprocess.CompletedProcess[str]:
        return subprocess.run(
            [sys.executable, str(SCRIPT), *args],
            text=True,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            check=False,
        )

    def create_signing_repo(
        self, curve: ec.EllipticCurve = ec.SECP256R1()
    ) -> Path:
        repo = self.root / "repo"
        repo.mkdir()
        private_key = ec.generate_private_key(curve)
        private_bytes = private_key.private_bytes(
            serialization.Encoding.PEM,
            serialization.PrivateFormat.PKCS8,
            serialization.NoEncryption(),
        )
        public_bytes = private_key.public_key().public_bytes(
            serialization.Encoding.DER,
            serialization.PublicFormat.SubjectPublicKeyInfo,
        )
        public_b64 = base64.b64encode(public_bytes).decode("ascii")
        private_path = repo / ".secure" / "amphion-license-private.pem"
        private_path.parent.mkdir()
        private_path.write_bytes(private_bytes)
        files = {
            "asr/android/gradle.properties": f"AMPHION_LICENSE_PUBLIC_KEY={public_b64}\n",
            "tts/android/gradle.properties": f"AMPHION_LICENSE_PUBLIC_KEY={public_b64}\n",
            "asr/harmony/sdk/src/main/ets/com/amphion/asr/License.ets": (
                f"const LICENSE_PUBLIC_KEY_B64: string = '{public_b64}';\n"
            ),
            "tts/harmony/sdk/src/main/ets/License.ets": (
                f"const LICENSE_PUBLIC_KEY_B64: string = '{public_b64}';\n"
            ),
        }
        for relative, content in files.items():
            path = repo / relative
            path.parent.mkdir(parents=True, exist_ok=True)
            path.write_text(content, encoding="utf-8")
        (repo / ".gitignore").write_text(".secure/\n", encoding="utf-8")
        subprocess.run(["git", "init", "-q"], cwd=repo, check=True)
        subprocess.run(
            ["git", "config", "user.email", "license-test@example.com"],
            cwd=repo,
            check=True,
        )
        subprocess.run(
            ["git", "config", "user.name", "License Test"], cwd=repo, check=True
        )
        subprocess.run(["git", "add", "."], cwd=repo, check=True)
        subprocess.run(
            ["git", "commit", "-q", "-m", "test: signing fixture"],
            cwd=repo,
            check=True,
        )
        return repo

    def test_plan_creates_non_sensitive_ready_receipt_for_valid_csv(self) -> None:
        source = self.input_dir / "devices.csv"
        source.write_text(
            "SN\n7GK0226310007121\n62Q0225C06020145\n",
            encoding="utf-8",
        )
        request = self.write_request(
            source.name,
            hashlib.sha256(source.read_bytes()).hexdigest(),
        )
        output = self.root / "plan.json"

        result = self.run_cli(
            "plan",
            "--request",
            str(request),
            "--input-dir",
            str(self.input_dir),
            "--out",
            str(output),
        )

        self.assertEqual(0, result.returncode, result.stderr)
        plan = json.loads(output.read_text(encoding="utf-8"))
        self.assertEqual("READY", plan["status"])
        self.assertEqual(2, plan["snSummary"]["nonEmptyCellCount"])
        self.assertEqual(2, plan["snSummary"]["uniqueCount"])
        self.assertEqual([], plan["warnings"])
        serialized = output.read_text(encoding="utf-8")
        self.assertNotIn("7GK0226310007121", serialized)
        self.assertNotIn("62Q0225C06020145", serialized)

    def test_plan_rejects_numeric_excel_sn_before_precision_can_be_lost(self) -> None:
        source = self.input_dir / "devices.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Sheet1"
        sheet.append(["SN"])
        sheet.append([1234567890123456])
        workbook.save(source)
        request = self.write_request(
            source.name,
            hashlib.sha256(source.read_bytes()).hexdigest(),
        )

        result = self.run_cli(
            "plan",
            "--request",
            str(request),
            "--input-dir",
            str(self.input_dir),
            "--out",
            str(self.root / "plan.json"),
        )

        self.assertNotEqual(0, result.returncode)
        self.assertIn("must be text", result.stderr)

    def test_plan_reads_explicit_columns_across_multiple_excel_sheets(self) -> None:
        source = self.input_dir / "devices.xlsx"
        workbook = Workbook()
        first = workbook.active
        first.title = "BatchA"
        first.append(["PrimarySN", "Ignore"])
        first.append(["7GK0226310007121", "not-selected"])
        second = workbook.create_sheet("BatchB")
        second.append(["StandardSN", "ProSN"])
        second.append(["62Q0225C06020145", "A2345678901234567"])
        workbook.save(source)
        request_path = self.write_request(
            source.name,
            hashlib.sha256(source.read_bytes()).hexdigest(),
        )
        request = json.loads(request_path.read_text(encoding="utf-8"))
        request["sources"][0]["sheets"] = [
            {"name": "BatchA", "columns": ["PrimarySN"]},
            {"name": "BatchB", "columns": ["StandardSN", "ProSN"]},
        ]
        request_path.write_text(json.dumps(request), encoding="utf-8")
        output = self.root / "plan.json"

        result = self.run_cli(
            "plan",
            "--request",
            str(request_path),
            "--input-dir",
            str(self.input_dir),
            "--out",
            str(output),
        )

        self.assertEqual(0, result.returncode, result.stderr)
        plan = json.loads(output.read_text(encoding="utf-8"))
        self.assertEqual(3, plan["snSummary"]["uniqueCount"])
        self.assertEqual({"16": 2, "17": 1}, plan["snSummary"]["lengthDistribution"])

    def test_plan_rejects_formula_missing_column_and_illegal_character(self) -> None:
        cases = (
            ("formula", "SN", "=A1", "must be text"),
            ("missing-column", "Other", "7GK0226310007121", "missing or ambiguous"),
            ("illegal-character", "SN", "SN-001", "outside [A-Z0-9]"),
        )
        for label, header, value, expected_error in cases:
            with self.subTest(label=label):
                source = self.input_dir / f"{label}.xlsx"
                workbook = Workbook()
                sheet = workbook.active
                sheet.title = "Sheet1"
                sheet.append([header])
                sheet.append([value])
                workbook.save(source)
                request = self.write_request(
                    source.name,
                    hashlib.sha256(source.read_bytes()).hexdigest(),
                )

                result = self.run_cli(
                    "plan",
                    "--request",
                    str(request),
                    "--input-dir",
                    str(self.input_dir),
                    "--out",
                    str(self.root / f"{label}.plan.json"),
                )

                self.assertNotEqual(0, result.returncode)
                self.assertIn(expected_error, result.stderr)

    def test_plan_requires_review_for_duplicate_outlier_and_hidden_excel_data(self) -> None:
        source = self.input_dir / "devices.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Sheet1"
        sheet.append(["SN"])
        sheet.append(["7GK0226310007121"])
        sheet.append([" 7gk0226310007121 "])
        sheet.append(["A2345678901234567"])
        sheet.row_dimensions[3].hidden = True
        sheet.column_dimensions["A"].hidden = True
        workbook.save(source)
        request = self.write_request(
            source.name,
            hashlib.sha256(source.read_bytes()).hexdigest(),
        )
        output = self.root / "plan.json"

        result = self.run_cli(
            "plan",
            "--request",
            str(request),
            "--input-dir",
            str(self.input_dir),
            "--out",
            str(output),
        )

        self.assertEqual(0, result.returncode, result.stderr)
        plan = json.loads(output.read_text(encoding="utf-8"))
        self.assertEqual("REVIEW_REQUIRED", plan["status"])
        self.assertEqual(
            {"DUPLICATE_SN", "SN_LENGTH_OUTLIER", "HIDDEN_ROWS", "HIDDEN_COLUMNS"},
            {warning["code"] for warning in plan["warnings"]},
        )
        self.assertEqual(3, plan["snSummary"]["nonEmptyCellCount"])
        self.assertEqual(2, plan["snSummary"]["uniqueCount"])

    def test_plan_rejects_implicit_perpetual_policy(self) -> None:
        source = self.input_dir / "devices.csv"
        source.write_text("SN\n7GK0226310007121\n", encoding="utf-8")
        request_path = self.write_request(
            source.name,
            hashlib.sha256(source.read_bytes()).hexdigest(),
        )
        request = json.loads(request_path.read_text(encoding="utf-8"))
        del request["policy"]["runtimeExpiry"]
        request_path.write_text(json.dumps(request), encoding="utf-8")

        result = self.run_cli(
            "plan",
            "--request",
            str(request_path),
            "--input-dir",
            str(self.input_dir),
            "--out",
            str(self.root / "plan.json"),
        )

        self.assertNotEqual(0, result.returncode)
        self.assertIn("runtimeExpiry", result.stderr)

    def test_plan_rejects_unknown_policy_fields_before_they_can_leak(self) -> None:
        source = self.input_dir / "devices.csv"
        source.write_text("SN\n7GK0226310007121\n", encoding="utf-8")
        request_path = self.write_request(
            source.name,
            hashlib.sha256(source.read_bytes()).hexdigest(),
        )
        request = json.loads(request_path.read_text(encoding="utf-8"))
        request["policy"]["authorizedDeviceHashes"] = ["sensitive-value"]
        request_path.write_text(json.dumps(request), encoding="utf-8")

        result = self.run_cli(
            "plan",
            "--request",
            str(request_path),
            "--input-dir",
            str(self.input_dir),
            "--out",
            str(self.root / "plan.json"),
        )

        self.assertNotEqual(0, result.returncode)
        self.assertIn("unknown policy fields", result.stderr)

    def test_plan_rejects_future_issue_date(self) -> None:
        source = self.input_dir / "devices.csv"
        source.write_text("SN\n7GK0226310007121\n", encoding="utf-8")
        request_path = self.write_request(
            source.name,
            hashlib.sha256(source.read_bytes()).hexdigest(),
        )
        request = json.loads(request_path.read_text(encoding="utf-8"))
        request["issuedAt"] = "2999-01-01"
        request_path.write_text(json.dumps(request), encoding="utf-8")

        result = self.run_cli(
            "plan",
            "--request",
            str(request_path),
            "--input-dir",
            str(self.input_dir),
            "--out",
            str(self.root / "plan.json"),
        )

        self.assertNotEqual(0, result.returncode)
        self.assertIn("future", result.stderr)

    def test_plan_accepts_utf8_text_with_one_sn_per_line(self) -> None:
        source = self.input_dir / "devices.txt"
        source.write_text(
            "# approved devices\n7GK0226310007121\n\n62Q0225C06020145\n",
            encoding="utf-8",
        )
        request = self.write_request(
            source.name,
            hashlib.sha256(source.read_bytes()).hexdigest(),
        )
        output = self.root / "plan.json"

        result = self.run_cli(
            "plan",
            "--request",
            str(request),
            "--input-dir",
            str(self.input_dir),
            "--out",
            str(output),
        )

        self.assertEqual(0, result.returncode, result.stderr)
        plan = json.loads(output.read_text(encoding="utf-8"))
        self.assertEqual(2, plan["snSummary"]["uniqueCount"])

    def test_plan_reports_counts_against_previous_full_snapshot(self) -> None:
        previous_input = self.root / "previous-input"
        previous_input.mkdir()
        previous_source = previous_input / "devices.csv"
        previous_source.write_text(
            "SN\n7GK0226310007121\n62Q0225C06020145\n",
            encoding="utf-8",
        )
        previous_request = self.write_request(
            previous_source.name,
            hashlib.sha256(previous_source.read_bytes()).hexdigest(),
        )
        previous_request = previous_request.rename(self.root / "previous-request.json")
        previous_payload = json.loads(previous_request.read_text(encoding="utf-8"))
        previous_payload["licenseId"] = "LIC-DQ-COMMERCIAL-000000"
        previous_request.write_text(json.dumps(previous_payload), encoding="utf-8")
        current_source = self.input_dir / "devices.csv"
        current_source.write_text(
            "SN\n7GK0226310007121\n62Q0225C06020145\n5JV0226415019854\n",
            encoding="utf-8",
        )
        current_request = self.write_request(
            current_source.name,
            hashlib.sha256(current_source.read_bytes()).hexdigest(),
        )
        current_payload = json.loads(current_request.read_text(encoding="utf-8"))
        current_payload["previousLicenseId"] = "LIC-DQ-COMMERCIAL-000000"
        current_request.write_text(json.dumps(current_payload), encoding="utf-8")
        output = self.root / "plan.json"

        result = self.run_cli(
            "plan",
            "--request",
            str(current_request),
            "--input-dir",
            str(self.input_dir),
            "--previous-request",
            str(previous_request),
            "--previous-input-dir",
            str(previous_input),
            "--out",
            str(output),
        )

        self.assertEqual(0, result.returncode, result.stderr)
        plan = json.loads(output.read_text(encoding="utf-8"))
        self.assertEqual(
            {
                "baseLicenseId": "LIC-DQ-COMMERCIAL-000000",
                "added": 1,
                "removed": 0,
                "unchanged": 2,
            },
            plan["diff"],
        )
        repo = self.create_signing_repo()
        issue_without_previous = self.run_cli(
            "issue",
            "--repo",
            str(repo),
            "--request",
            str(current_request),
            "--plan",
            str(output),
            "--input-dir",
            str(self.input_dir),
            "--operator",
            "issuer@example.com",
            "--out-dir",
            str(self.root / "output"),
        )
        self.assertNotEqual(0, issue_without_previous.returncode)
        self.assertIn("previous", issue_without_previous.stderr)

    def test_issue_creates_customer_zip_without_plaintext_sn(self) -> None:
        source = self.input_dir / "devices.csv"
        source.write_text(
            "SN\n7GK0226310007121\n62Q0225C06020145\n",
            encoding="utf-8",
        )
        request = self.write_request(
            source.name,
            hashlib.sha256(source.read_bytes()).hexdigest(),
        )
        plan_path = self.root / "plan.json"
        plan_result = self.run_cli(
            "plan",
            "--request",
            str(request),
            "--input-dir",
            str(self.input_dir),
            "--out",
            str(plan_path),
        )
        self.assertEqual(0, plan_result.returncode, plan_result.stderr)
        repo = self.create_signing_repo()
        output_dir = self.root / "output"

        result = self.run_cli(
            "issue",
            "--repo",
            str(repo),
            "--request",
            str(request),
            "--plan",
            str(plan_path),
            "--input-dir",
            str(self.input_dir),
            "--operator",
            "tester@example.com",
            "--out-dir",
            str(output_dir),
        )

        self.assertEqual(0, result.returncode, result.stderr)
        zip_path = output_dir / "DEL-DQ-COMMERCIAL-20260730-001.zip"
        self.assertTrue(zip_path.is_file())
        with zipfile.ZipFile(zip_path) as archive:
            names = set(archive.namelist())
            root = "DEL-DQ-COMMERCIAL-20260730-001/"
            self.assertEqual(
                {
                    root + "amphion-license.lic",
                    root + "README.md",
                    root + "LICENSE_MANIFEST.json",
                    root + "LICENSE_VERIFICATION.json",
                    root + "LICENSE_VERIFICATION.md",
                    root + "SHA256SUMS.txt",
                },
                names,
            )
            envelope = json.loads(archive.read(root + "amphion-license.lic"))
            claims = json.loads(base64.b64decode(envelope["payload_b64"]))
            self.assertEqual(2, len(claims["authorizedDeviceHashes"]))
            combined = b"".join(archive.read(name) for name in names)
            self.assertNotIn(b"7GK0226310007121", combined)
            self.assertNotIn(b"62Q0225C06020145", combined)
        receipt = json.loads(
            (output_dir / "DEL-DQ-COMMERCIAL-20260730-001.issuance.json").read_text(
                encoding="utf-8"
            )
        )
        self.assertTrue(receipt["production"])
        self.assertEqual("tester@example.com", receipt["operator"])

    def test_issue_rejects_changed_inputs_and_modified_plan(self) -> None:
        source = self.input_dir / "devices.csv"
        source.write_text("SN\n7GK0226310007121\n", encoding="utf-8")
        request = self.write_request(
            source.name,
            hashlib.sha256(source.read_bytes()).hexdigest(),
        )
        plan_path = self.root / "plan.json"
        self.assertEqual(
            0,
            self.run_cli(
                "plan",
                "--request",
                str(request),
                "--input-dir",
                str(self.input_dir),
                "--out",
                str(plan_path),
            ).returncode,
        )
        repo = self.create_signing_repo()
        common = (
            "issue",
            "--repo",
            str(repo),
            "--request",
            str(request),
            "--plan",
            str(plan_path),
            "--input-dir",
            str(self.input_dir),
            "--operator",
            "issuer@example.com",
            "--out-dir",
            str(self.root / "output"),
        )
        source.write_text("SN\n62Q0225C06020145\n", encoding="utf-8")
        changed_request = json.loads(request.read_text(encoding="utf-8"))
        changed_request["sources"][0]["sha256"] = hashlib.sha256(
            source.read_bytes()
        ).hexdigest()
        request.write_text(json.dumps(changed_request), encoding="utf-8")

        changed = self.run_cli(*common)
        self.assertNotEqual(0, changed.returncode)
        self.assertIn("approved plan", changed.stderr)

        source.write_text("SN\n7GK0226310007121\n", encoding="utf-8")
        original_request = json.loads(request.read_text(encoding="utf-8"))
        original_request["sources"][0]["sha256"] = hashlib.sha256(
            source.read_bytes()
        ).hexdigest()
        request.write_text(json.dumps(original_request), encoding="utf-8")
        modified_plan = json.loads(plan_path.read_text(encoding="utf-8"))
        modified_plan["snSummary"]["uniqueCount"] = 999
        plan_path.write_text(json.dumps(modified_plan), encoding="utf-8")

        modified = self.run_cli(*common)
        self.assertNotEqual(0, modified.returncode)
        self.assertIn("plan receipt", modified.stderr)

    def test_issue_requires_all_plan_warning_acknowledgements(self) -> None:
        source = self.input_dir / "devices.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Sheet1"
        sheet.append(["SN"])
        sheet.append(["7GK0226310007121"])
        sheet.append(["7gk0226310007121"])
        workbook.save(source)
        request = self.write_request(
            source.name,
            hashlib.sha256(source.read_bytes()).hexdigest(),
        )
        plan_path = self.root / "plan.json"
        self.assertEqual(
            0,
            self.run_cli(
                "plan",
                "--request",
                str(request),
                "--input-dir",
                str(self.input_dir),
                "--out",
                str(plan_path),
            ).returncode,
        )
        repo = self.create_signing_repo()

        result = self.run_cli(
            "issue",
            "--repo",
            str(repo),
            "--request",
            str(request),
            "--plan",
            str(plan_path),
            "--input-dir",
            str(self.input_dir),
            "--operator",
            "issuer@example.com",
            "--out-dir",
            str(self.root / "output"),
        )

        self.assertNotEqual(0, result.returncode)
        self.assertIn("DUPLICATE_SN", result.stderr)

    def test_issue_dirty_worktree_is_non_production_only(self) -> None:
        source = self.input_dir / "devices.csv"
        source.write_text("SN\n7GK0226310007121\n", encoding="utf-8")
        request = self.write_request(
            source.name,
            hashlib.sha256(source.read_bytes()).hexdigest(),
        )
        plan_path = self.root / "plan.json"
        self.assertEqual(
            0,
            self.run_cli(
                "plan",
                "--request",
                str(request),
                "--input-dir",
                str(self.input_dir),
                "--out",
                str(plan_path),
            ).returncode,
        )
        repo = self.create_signing_repo()
        (repo / "dirty.txt").write_text("dirty", encoding="utf-8")
        common = (
            "issue",
            "--repo",
            str(repo),
            "--request",
            str(request),
            "--plan",
            str(plan_path),
            "--input-dir",
            str(self.input_dir),
            "--operator",
            "issuer@example.com",
            "--out-dir",
            str(self.root / "output"),
        )

        rejected = self.run_cli(*common)
        self.assertNotEqual(0, rejected.returncode)
        self.assertIn("clean", rejected.stderr)
        allowed = self.run_cli(*common, "--allow-dirty")
        self.assertEqual(0, allowed.returncode, allowed.stderr)
        receipt = json.loads(
            (self.root / "output/DEL-DQ-COMMERCIAL-20260730-001.issuance.json").read_text(
                encoding="utf-8"
            )
        )
        self.assertFalse(receipt["production"])

    def test_issue_rejects_non_p256_signing_key(self) -> None:
        source = self.input_dir / "devices.csv"
        source.write_text("SN\n7GK0226310007121\n", encoding="utf-8")
        request = self.write_request(
            source.name,
            hashlib.sha256(source.read_bytes()).hexdigest(),
        )
        plan_path = self.root / "plan.json"
        self.assertEqual(
            0,
            self.run_cli(
                "plan",
                "--request",
                str(request),
                "--input-dir",
                str(self.input_dir),
                "--out",
                str(plan_path),
            ).returncode,
        )
        repo = self.create_signing_repo(ec.SECP384R1())

        result = self.run_cli(
            "issue",
            "--repo",
            str(repo),
            "--request",
            str(request),
            "--plan",
            str(plan_path),
            "--input-dir",
            str(self.input_dir),
            "--operator",
            "issuer@example.com",
            "--out-dir",
            str(self.root / "output"),
        )

        self.assertNotEqual(0, result.returncode)
        self.assertIn("P-256", result.stderr)

    def test_issue_rejects_when_one_sdk_public_key_differs(self) -> None:
        source = self.input_dir / "devices.csv"
        source.write_text("SN\n7GK0226310007121\n", encoding="utf-8")
        request = self.write_request(
            source.name,
            hashlib.sha256(source.read_bytes()).hexdigest(),
        )
        plan_path = self.root / "plan.json"
        self.assertEqual(
            0,
            self.run_cli(
                "plan",
                "--request",
                str(request),
                "--input-dir",
                str(self.input_dir),
                "--out",
                str(plan_path),
            ).returncode,
        )
        repo = self.create_signing_repo()
        other_key = ec.generate_private_key(ec.SECP256R1()).public_key().public_bytes(
            serialization.Encoding.DER,
            serialization.PublicFormat.SubjectPublicKeyInfo,
        )
        (repo / "tts/android/gradle.properties").write_text(
            "AMPHION_LICENSE_PUBLIC_KEY="
            + base64.b64encode(other_key).decode("ascii")
            + "\n",
            encoding="utf-8",
        )
        subprocess.run(["git", "add", "."], cwd=repo, check=True)
        subprocess.run(
            ["git", "commit", "-q", "-m", "test: mismatched public key"],
            cwd=repo,
            check=True,
        )

        result = self.run_cli(
            "issue",
            "--repo",
            str(repo),
            "--request",
            str(request),
            "--plan",
            str(plan_path),
            "--input-dir",
            str(self.input_dir),
            "--operator",
            "issuer@example.com",
            "--out-dir",
            str(self.root / "output"),
        )

        self.assertNotEqual(0, result.returncode)
        self.assertIn("four SDK", result.stderr)

    def test_verify_reopens_final_zip_and_writes_external_pass_receipt(self) -> None:
        source = self.input_dir / "devices.csv"
        source.write_text(
            "SN\n7GK0226310007121\n62Q0225C06020145\n",
            encoding="utf-8",
        )
        request = self.write_request(
            source.name,
            hashlib.sha256(source.read_bytes()).hexdigest(),
        )
        plan_path = self.root / "plan.json"
        self.assertEqual(
            0,
            self.run_cli(
                "plan",
                "--request",
                str(request),
                "--input-dir",
                str(self.input_dir),
                "--out",
                str(plan_path),
            ).returncode,
        )
        repo = self.create_signing_repo()
        output_dir = self.root / "output"
        issue_result = self.run_cli(
            "issue",
            "--repo",
            str(repo),
            "--request",
            str(request),
            "--plan",
            str(plan_path),
            "--input-dir",
            str(self.input_dir),
            "--operator",
            "issuer@example.com",
            "--out-dir",
            str(output_dir),
        )
        self.assertEqual(0, issue_result.returncode, issue_result.stderr)
        zip_path = output_dir / "DEL-DQ-COMMERCIAL-20260730-001.zip"
        prefix = output_dir / "DEL-DQ-COMMERCIAL-20260730-001.zip"

        renamed_zip = output_dir / "renamed.zip"
        shutil.copyfile(zip_path, renamed_zip)
        renamed = self.run_cli(
            "verify",
            "--repo",
            str(repo),
            "--request",
            str(request),
            "--plan",
            str(plan_path),
            "--input-dir",
            str(self.input_dir),
            "--zip",
            str(renamed_zip),
            "--operator",
            "verifier@example.com",
            "--out-prefix",
            str(output_dir / "renamed.zip"),
        )
        self.assertNotEqual(0, renamed.returncode)
        self.assertIn("file name", renamed.stderr)

        result = self.run_cli(
            "verify",
            "--repo",
            str(repo),
            "--request",
            str(request),
            "--plan",
            str(plan_path),
            "--input-dir",
            str(self.input_dir),
            "--zip",
            str(zip_path),
            "--operator",
            "verifier@example.com",
            "--out-prefix",
            str(prefix),
        )

        self.assertEqual(0, result.returncode, result.stderr)
        receipt_path = Path(str(prefix) + ".verification.json")
        report_path = Path(str(prefix) + ".verification.md")
        receipt = json.loads(receipt_path.read_text(encoding="utf-8"))
        self.assertEqual("PASS", receipt["status"])
        self.assertEqual(
            hashlib.sha256(zip_path.read_bytes()).hexdigest(), receipt["zipSha256"]
        )
        self.assertEqual("verifier@example.com", receipt["operator"])
        self.assertTrue(report_path.is_file())

        root = "DEL-DQ-COMMERCIAL-20260730-001/"
        with zipfile.ZipFile(zip_path) as archive:
            readme = archive.read(root + "README.md")
        with warnings.catch_warnings():
            warnings.simplefilter("ignore", UserWarning)
            with zipfile.ZipFile(zip_path, "a") as archive:
                archive.writestr(root + "README.md", readme)
        duplicate = self.run_cli(
            "verify",
            "--repo",
            str(repo),
            "--request",
            str(request),
            "--plan",
            str(plan_path),
            "--input-dir",
            str(self.input_dir),
            "--zip",
            str(zip_path),
            "--operator",
            "verifier@example.com",
            "--out-prefix",
            str(output_dir / "duplicate"),
        )
        self.assertNotEqual(0, duplicate.returncode)
        self.assertIn("duplicate ZIP", duplicate.stderr)

    def test_verify_requires_p256_even_when_all_sdk_keys_match(self) -> None:
        source = self.input_dir / "devices.csv"
        source.write_text("SN\n7GK0226310007121\n", encoding="utf-8")
        request = self.write_request(
            source.name,
            hashlib.sha256(source.read_bytes()).hexdigest(),
        )
        plan_path = self.root / "plan.json"
        self.assertEqual(
            0,
            self.run_cli(
                "plan",
                "--request",
                str(request),
                "--input-dir",
                str(self.input_dir),
                "--out",
                str(plan_path),
            ).returncode,
        )
        repo = self.create_signing_repo()
        output_dir = self.root / "output"
        self.assertEqual(
            0,
            self.run_cli(
                "issue",
                "--repo",
                str(repo),
                "--request",
                str(request),
                "--plan",
                str(plan_path),
                "--input-dir",
                str(self.input_dir),
                "--operator",
                "issuer@example.com",
                "--out-dir",
                str(output_dir),
            ).returncode,
        )
        public_der = ec.generate_private_key(ec.SECP384R1()).public_key().public_bytes(
            serialization.Encoding.DER,
            serialization.PublicFormat.SubjectPublicKeyInfo,
        )
        public_b64 = base64.b64encode(public_der).decode("ascii")
        replacements = {
            "asr/android/gradle.properties": f"AMPHION_LICENSE_PUBLIC_KEY={public_b64}\n",
            "tts/android/gradle.properties": f"AMPHION_LICENSE_PUBLIC_KEY={public_b64}\n",
            "asr/harmony/sdk/src/main/ets/com/amphion/asr/License.ets": (
                f"const LICENSE_PUBLIC_KEY_B64: string = '{public_b64}';\n"
            ),
            "tts/harmony/sdk/src/main/ets/License.ets": (
                f"const LICENSE_PUBLIC_KEY_B64: string = '{public_b64}';\n"
            ),
        }
        for relative, content in replacements.items():
            (repo / relative).write_text(content, encoding="utf-8")

        result = self.run_cli(
            "verify",
            "--repo",
            str(repo),
            "--request",
            str(request),
            "--plan",
            str(plan_path),
            "--input-dir",
            str(self.input_dir),
            "--zip",
            str(output_dir / "DEL-DQ-COMMERCIAL-20260730-001.zip"),
            "--operator",
            "verifier@example.com",
            "--out-prefix",
            str(output_dir / "p384"),
        )

        self.assertNotEqual(0, result.returncode)
        self.assertIn("P-256", result.stderr)

    def test_verify_rejects_tampered_zip_member(self) -> None:
        source = self.input_dir / "devices.csv"
        source.write_text("SN\n7GK0226310007121\n", encoding="utf-8")
        request = self.write_request(
            source.name,
            hashlib.sha256(source.read_bytes()).hexdigest(),
        )
        plan_path = self.root / "plan.json"
        self.assertEqual(
            0,
            self.run_cli(
                "plan",
                "--request",
                str(request),
                "--input-dir",
                str(self.input_dir),
                "--out",
                str(plan_path),
            ).returncode,
        )
        repo = self.create_signing_repo()
        output_dir = self.root / "output"
        self.assertEqual(
            0,
            self.run_cli(
                "issue",
                "--repo",
                str(repo),
                "--request",
                str(request),
                "--plan",
                str(plan_path),
                "--input-dir",
                str(self.input_dir),
                "--operator",
                "issuer@example.com",
                "--out-dir",
                str(output_dir),
            ).returncode,
        )
        zip_path = output_dir / "DEL-DQ-COMMERCIAL-20260730-001.zip"
        root = "DEL-DQ-COMMERCIAL-20260730-001/"
        with zipfile.ZipFile(zip_path) as source_zip:
            members = {
                name: source_zip.read(name) for name in source_zip.namelist()
            }
        members[root + "README.md"] += b"customer-facing text changed\n"
        checksum_members = {
            name.removeprefix(root): content
            for name, content in members.items()
            if name != root + "SHA256SUMS.txt"
        }
        members[root + "SHA256SUMS.txt"] = "".join(
            f"{hashlib.sha256(content).hexdigest()}  {name}\n"
            for name, content in sorted(checksum_members.items())
        ).encode("utf-8")
        with zipfile.ZipFile(zip_path, "w") as target_zip:
            for name, content in members.items():
                target_zip.writestr(name, content)

        result = self.run_cli(
            "verify",
            "--repo",
            str(repo),
            "--request",
            str(request),
            "--plan",
            str(plan_path),
            "--input-dir",
            str(self.input_dir),
            "--zip",
            str(zip_path),
            "--operator",
            "verifier@example.com",
            "--out-prefix",
            str(output_dir / zip_path.name),
        )

        self.assertNotEqual(0, result.returncode)
        self.assertIn("README.md", result.stderr)

    def test_record_appends_verified_delivery_without_sensitive_sn_metadata(self) -> None:
        source = self.input_dir / "devices.csv"
        source.write_text(
            "SN\n7GK0226310007121\n62Q0225C06020145\n",
            encoding="utf-8",
        )
        request = self.write_request(
            source.name,
            hashlib.sha256(source.read_bytes()).hexdigest(),
        )
        plan_path = self.root / "plan.json"
        self.assertEqual(
            0,
            self.run_cli(
                "plan",
                "--request",
                str(request),
                "--input-dir",
                str(self.input_dir),
                "--out",
                str(plan_path),
            ).returncode,
        )
        repo = self.create_signing_repo()
        output_dir = self.root / "output"
        issue_result = self.run_cli(
            "issue",
            "--repo",
            str(repo),
            "--request",
            str(request),
            "--plan",
            str(plan_path),
            "--input-dir",
            str(self.input_dir),
            "--operator",
            "issuer@example.com",
            "--out-dir",
            str(output_dir),
        )
        self.assertEqual(0, issue_result.returncode, issue_result.stderr)
        zip_path = output_dir / "DEL-DQ-COMMERCIAL-20260730-001.zip"
        prefix = output_dir / zip_path.name
        verify_result = self.run_cli(
            "verify",
            "--repo",
            str(repo),
            "--request",
            str(request),
            "--plan",
            str(plan_path),
            "--input-dir",
            str(self.input_dir),
            "--zip",
            str(zip_path),
            "--operator",
            "verifier@example.com",
            "--out-prefix",
            str(prefix),
        )
        self.assertEqual(0, verify_result.returncode, verify_result.stderr)
        history = repo / "delivery" / "license-delivery-history.json"
        verification_path = Path(str(prefix) + ".verification.json")
        verification_payload = json.loads(verification_path.read_text(encoding="utf-8"))
        verification_payload["status"] = "FAIL"
        verification_path.write_text(json.dumps(verification_payload), encoding="utf-8")

        rejected = self.run_cli(
            "record",
            "--repo",
            str(repo),
            "--plan",
            str(plan_path),
            "--zip",
            str(zip_path),
            "--issuance",
            str(output_dir / "DEL-DQ-COMMERCIAL-20260730-001.issuance.json"),
            "--verification",
            str(verification_path),
            "--operator",
            "delivery@example.com",
            "--delivered-at",
            "2026-07-30",
        )
        self.assertNotEqual(0, rejected.returncode)
        self.assertIn("not PASS", rejected.stderr)
        verification_payload["status"] = "PASS"
        verification_path.write_text(json.dumps(verification_payload), encoding="utf-8")

        result = self.run_cli(
            "record",
            "--repo",
            str(repo),
            "--plan",
            str(plan_path),
            "--zip",
            str(zip_path),
            "--issuance",
            str(output_dir / "DEL-DQ-COMMERCIAL-20260730-001.issuance.json"),
            "--verification",
            str(verification_path),
            "--operator",
            "delivery@example.com",
            "--delivered-at",
            "2026-07-30",
        )

        self.assertEqual(0, result.returncode, result.stderr)
        payload = json.loads(history.read_text(encoding="utf-8"))
        self.assertEqual(1, len(payload["deliveries"]))
        entry = payload["deliveries"][0]
        self.assertEqual("delivered", entry["status"])
        self.assertEqual("delivery@example.com", entry["operator"])
        self.assertEqual(2, entry["uniqueSnCount"])
        serialized = history.read_text(encoding="utf-8")
        self.assertNotIn("snSetDigest", serialized)
        self.assertNotIn("sourceSha256", serialized)
        self.assertNotIn("7GK0226310007121", serialized)
        self.assertEqual(
            hashlib.sha256(plan_path.read_bytes()).hexdigest(),
            entry["planReceiptSha256"],
        )

        subprocess.run(["git", "add", str(history)], cwd=repo, check=True)
        subprocess.run(
            ["git", "commit", "-q", "-m", "test: record delivery"],
            cwd=repo,
            check=True,
        )
        duplicate = self.run_cli(
            "record",
            "--repo",
            str(repo),
            "--plan",
            str(plan_path),
            "--zip",
            str(zip_path),
            "--issuance",
            str(output_dir / "DEL-DQ-COMMERCIAL-20260730-001.issuance.json"),
            "--verification",
            str(verification_path),
            "--operator",
            "delivery@example.com",
            "--delivered-at",
            "2026-07-30",
        )
        self.assertNotEqual(0, duplicate.returncode)
        self.assertIn("already contains", duplicate.stderr)


if __name__ == "__main__":
    unittest.main()
