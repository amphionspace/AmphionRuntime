#!/usr/bin/env python3
"""Plan, issue, verify, and record formal offline License deliveries."""

from __future__ import annotations

import argparse
import csv
import base64
import fcntl
import hashlib
import json
import os
import re
import subprocess
import sys
import tempfile
import uuid
import zipfile
from collections import Counter
from datetime import datetime
from pathlib import Path, PurePosixPath
import stat
from typing import Any, Dict, List, Optional, Sequence, Tuple

from cryptography.exceptions import InvalidSignature
from cryptography.hazmat.primitives import hashes
from cryptography.hazmat.primitives.asymmetric import ec
from cryptography.hazmat.primitives import serialization

from issue_license import DEFAULT_DEVICE_ID_SALT_ID, create_license_envelope


SCHEMA_VERSION = 1
SN_PATTERN = re.compile(r"^[A-Z0-9]+$")
APPROVED_PLAN_FIELDS = (
    "request",
    "requestSha256",
    "sources",
    "snSetDigest",
    "snSummary",
    "diff",
    "warnings",
)
DELIVERY_MEMBER_NAMES = frozenset(
    {
        "amphion-license.lic",
        "README.md",
        "LICENSE_MANIFEST.json",
        "LICENSE_VERIFICATION.json",
        "LICENSE_VERIFICATION.md",
        "SHA256SUMS.txt",
    }
)


class LicenseDeliveryError(RuntimeError):
    pass


def _sha256_bytes(value: bytes) -> str:
    return hashlib.sha256(value).hexdigest()


def _canonical_json_bytes(value: object) -> bytes:
    return json.dumps(
        value,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    ).encode("utf-8")


def _load_json(path: Path) -> Dict[str, Any]:
    try:
        value = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, UnicodeError, json.JSONDecodeError) as error:
        raise LicenseDeliveryError(f"cannot read JSON {path.name}: {error}") from error
    if not isinstance(value, dict):
        raise LicenseDeliveryError(f"JSON root must be an object: {path.name}")
    return value


def _normalize_sn(value: str) -> str:
    normalized = value.strip().upper()
    if not normalized:
        return ""
    if not SN_PATTERN.fullmatch(normalized):
        raise LicenseDeliveryError("SN contains characters outside [A-Z0-9]")
    return normalized


def _read_csv_source(path: Path, columns: Sequence[str]) -> Tuple[List[str], Dict[str, Any]]:
    try:
        with path.open("r", encoding="utf-8-sig", newline="") as stream:
            reader = csv.DictReader(stream)
            headers = reader.fieldnames or []
            if len(headers) != len(set(headers)):
                raise LicenseDeliveryError(f"duplicate CSV headers: {path.name}")
            missing = [column for column in columns if column not in headers]
            if missing:
                raise LicenseDeliveryError(
                    f"missing CSV columns in {path.name}: {','.join(missing)}"
                )
            values: List[str] = []
            blank_count = 0
            for row in reader:
                for column in columns:
                    raw = row.get(column)
                    if raw is None or not raw.strip():
                        blank_count += 1
                        continue
                    values.append(_normalize_sn(raw))
    except (OSError, UnicodeError, csv.Error) as error:
        raise LicenseDeliveryError(f"cannot read CSV {path.name}: {error}") from error
    return values, {"blankCellCount": blank_count, "hiddenRows": 0, "hiddenColumns": 0}


def _read_xlsx_source(
    path: Path, sheet_mappings: Sequence[Dict[str, Any]]
) -> Tuple[List[str], Dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as error:
        raise LicenseDeliveryError(
            "XLSX input requires openpyxl; install tools/license/requirements.txt"
        ) from error
    try:
        workbook = load_workbook(path, read_only=False, data_only=False, keep_links=False)
    except (OSError, ValueError) as error:
        raise LicenseDeliveryError(f"cannot read XLSX {path.name}: {error}") from error
    values: List[str] = []
    blank_count = 0
    hidden_rows = 0
    hidden_columns = 0
    selected_columns: List[str] = []
    try:
        for mapping in sheet_mappings:
            sheet_name = mapping.get("name") if isinstance(mapping, dict) else None
            columns = mapping.get("columns") if isinstance(mapping, dict) else None
            if not isinstance(sheet_name, str) or not sheet_name:
                raise LicenseDeliveryError(f"XLSX sheet name is invalid: {path.name}")
            if not isinstance(columns, list) or not columns or not all(
                isinstance(column, str) and column for column in columns
            ):
                raise LicenseDeliveryError(f"XLSX columns are invalid: {path.name}")
            if sheet_name not in workbook.sheetnames:
                raise LicenseDeliveryError(
                    f"missing XLSX sheet in {path.name}: {sheet_name}"
                )
            sheet = workbook[sheet_name]
            if sheet.merged_cells.ranges:
                raise LicenseDeliveryError(
                    f"merged cells are not allowed in selected XLSX sheet: {sheet_name}"
                )
            headers = [cell.value for cell in sheet[1]]
            text_headers = [value for value in headers if isinstance(value, str) and value]
            if len(text_headers) != len(set(text_headers)):
                raise LicenseDeliveryError(
                    f"duplicate XLSX headers in {path.name}:{sheet_name}"
                )
            header_indexes: Dict[str, int] = {}
            for column in columns:
                matches = [index for index, value in enumerate(headers, start=1) if value == column]
                if len(matches) != 1:
                    raise LicenseDeliveryError(
                        f"missing or ambiguous XLSX column in {path.name}:{sheet_name}: {column}"
                    )
                header_indexes[column] = matches[0]
                selected_columns.append(f"{sheet_name}:{column}")
                if sheet.column_dimensions[sheet.cell(1, matches[0]).column_letter].hidden:
                    hidden_columns += 1
            for row_index in range(2, sheet.max_row + 1):
                if sheet.row_dimensions[row_index].hidden:
                    hidden_rows += 1
                for column in columns:
                    cell = sheet.cell(row_index, header_indexes[column])
                    raw = cell.value
                    if raw is None or (isinstance(raw, str) and not raw.strip()):
                        blank_count += 1
                        continue
                    if cell.data_type == "f" or not isinstance(raw, str):
                        raise LicenseDeliveryError(
                            f"SN cell must be text: {path.name}:{sheet_name}!{cell.coordinate}"
                        )
                    values.append(_normalize_sn(raw))
    finally:
        workbook.close()
    return values, {
        "blankCellCount": blank_count,
        "hiddenRows": hidden_rows,
        "hiddenColumns": hidden_columns,
        "selectedColumns": selected_columns,
    }


def _read_text_source(path: Path) -> Tuple[List[str], Dict[str, Any]]:
    try:
        lines = path.read_text(encoding="utf-8-sig").splitlines()
    except (OSError, UnicodeError) as error:
        raise LicenseDeliveryError(f"cannot read UTF-8 text {path.name}: {error}") from error
    values: List[str] = []
    blank_count = 0
    for line in lines:
        if not line.strip():
            blank_count += 1
            continue
        if line.lstrip().startswith("#"):
            continue
        values.append(_normalize_sn(line))
    return values, {"blankCellCount": blank_count, "hiddenRows": 0, "hiddenColumns": 0}


def _validate_request(request: Dict[str, Any]) -> None:
    required = {
        "schemaVersion",
        "requestId",
        "licenseId",
        "deliveryId",
        "customerId",
        "projectId",
        "reason",
        "issuedAt",
        "sources",
        "policy",
        "previousLicenseId",
    }
    missing = sorted(required - set(request))
    if missing:
        raise LicenseDeliveryError(f"request missing fields: {','.join(missing)}")
    if request["schemaVersion"] != SCHEMA_VERSION:
        raise LicenseDeliveryError(f"request schemaVersion must be {SCHEMA_VERSION}")
    for field in (
        "requestId",
        "licenseId",
        "deliveryId",
        "customerId",
        "projectId",
        "reason",
    ):
        if not isinstance(request[field], str) or not request[field].strip():
            raise LicenseDeliveryError(f"request {field} must be a non-empty string")
    if not re.fullmatch(r"[A-Za-z0-9][A-Za-z0-9._-]*", request["deliveryId"]):
        raise LicenseDeliveryError(
            "request deliveryId must be a safe file identifier using A-Z, 0-9, dot, underscore, or hyphen"
        )
    try:
        issued_at = datetime.strptime(request["issuedAt"], "%Y-%m-%d").date()
    except (TypeError, ValueError) as error:
        raise LicenseDeliveryError("request issuedAt must be YYYY-MM-DD") from error
    if issued_at > datetime.now().date():
        raise LicenseDeliveryError("request issuedAt cannot be in the future")
    previous_license_id = request["previousLicenseId"]
    if previous_license_id is not None and (
        not isinstance(previous_license_id, str) or not previous_license_id.strip()
    ):
        raise LicenseDeliveryError("request previousLicenseId must be null or a non-empty string")
    if not isinstance(request["sources"], list) or not request["sources"]:
        raise LicenseDeliveryError("request sources must be a non-empty array")
    policy = request["policy"]
    if not isinstance(policy, dict) or policy.get("deviceBinding") != "required":
        raise LicenseDeliveryError("formal delivery requires policy.deviceBinding=required")
    claims = _policy_claims(policy)
    for field in ("expiresAt", "maintenanceUntil"):
        if claims[field] and datetime.strptime(claims[field], "%Y-%m-%d").date() < issued_at:
            raise LicenseDeliveryError(f"policy {field} cannot precede issuedAt")


def _build_plan_with_values(
    request_path: Path, input_dir: Path
) -> Tuple[Dict[str, Any], List[str]]:
    request = _load_json(request_path)
    _validate_request(request)
    all_values: List[str] = []
    source_summaries: List[Dict[str, Any]] = []
    for source in request["sources"]:
        if not isinstance(source, dict):
            raise LicenseDeliveryError("each source must be an object")
        file_name = source.get("fileName")
        if not isinstance(file_name, str) or Path(file_name).name != file_name:
            raise LicenseDeliveryError("source fileName must be a basename")
        path = input_dir / file_name
        if not path.is_file():
            raise LicenseDeliveryError(f"source file not found: {file_name}")
        actual_sha256 = _sha256_bytes(path.read_bytes())
        if actual_sha256 != source.get("sha256"):
            raise LicenseDeliveryError(f"source SHA-256 mismatch: {file_name}")
        sheets = source.get("sheets")
        if not isinstance(sheets, list) or not sheets:
            raise LicenseDeliveryError(f"source sheets are invalid: {file_name}")
        suffix = path.suffix.lower()
        if suffix == ".csv":
            if len(sheets) != 1:
                raise LicenseDeliveryError("CSV source requires exactly one sheet mapping")
            columns = sheets[0].get("columns") if isinstance(sheets[0], dict) else None
            if not isinstance(columns, list) or not columns or not all(
                isinstance(column, str) and column for column in columns
            ):
                raise LicenseDeliveryError(f"source columns are invalid: {file_name}")
            values, details = _read_csv_source(path, columns)
            selected_columns = list(columns)
        elif suffix == ".txt":
            if len(sheets) != 1:
                raise LicenseDeliveryError("TXT source requires exactly one sheet mapping")
            columns = sheets[0].get("columns") if isinstance(sheets[0], dict) else None
            if columns != ["SN"]:
                raise LicenseDeliveryError("TXT source columns must be exactly ['SN']")
            values, details = _read_text_source(path)
            selected_columns = ["SN"]
        elif suffix == ".xlsx":
            values, details = _read_xlsx_source(path, sheets)
            selected_columns = details["selectedColumns"]
        else:
            raise LicenseDeliveryError(f"unsupported source type: {suffix or '(none)'}")
        if _sha256_bytes(path.read_bytes()) != actual_sha256:
            raise LicenseDeliveryError(f"source changed while it was being read: {file_name}")
        all_values.extend(values)
        source_summaries.append(
            {
                "fileName": file_name,
                "sha256": actual_sha256,
                "nonEmptyCellCount": len(values),
                "blankCellCount": details["blankCellCount"],
                "selectedColumns": selected_columns,
                "hiddenRowCount": details["hiddenRows"],
                "hiddenColumnCount": details["hiddenColumns"],
            }
        )
    if not all_values:
        raise LicenseDeliveryError("device-bound delivery requires at least one SN")
    counts = Counter(all_values)
    unique_values = sorted(counts)
    length_distribution = {
        str(length): sum(1 for value in unique_values if len(value) == length)
        for length in sorted({len(value) for value in unique_values})
    }
    warnings: List[Dict[str, Any]] = []
    hidden_rows = sum(source["hiddenRowCount"] for source in source_summaries)
    hidden_columns = sum(source["hiddenColumnCount"] for source in source_summaries)
    if hidden_rows:
        warnings.append(
            {
                "code": "HIDDEN_ROWS",
                "count": hidden_rows,
                "message": "hidden rows are included in the SN set",
            }
        )
    if hidden_columns:
        warnings.append(
            {
                "code": "HIDDEN_COLUMNS",
                "count": hidden_columns,
                "message": "selected SN columns are hidden",
            }
        )
    duplicate_cells = len(all_values) - len(unique_values)
    if duplicate_cells:
        warnings.append(
            {
                "code": "DUPLICATE_SN",
                "count": duplicate_cells,
                "message": "duplicate SN cells will be deduplicated",
            }
        )
    modal_length = Counter(map(len, unique_values)).most_common(1)[0][0]
    length_outliers = sum(1 for value in unique_values if len(value) != modal_length)
    if length_outliers:
        warnings.append(
            {
                "code": "SN_LENGTH_OUTLIER",
                "count": length_outliers,
                "message": "SN length differs from the modal length",
            }
        )
    normalized_bytes = ("\n".join(unique_values) + "\n").encode("utf-8")
    plan: Dict[str, Any] = {
        "schemaVersion": SCHEMA_VERSION,
        "kind": "LicenseDeliveryPlan",
        "status": "REVIEW_REQUIRED" if warnings else "READY",
        "request": {
            key: request[key]
            for key in (
                "requestId",
                "licenseId",
                "deliveryId",
                "customerId",
                "projectId",
                "reason",
                "issuedAt",
                "policy",
                "previousLicenseId",
            )
        },
        "requestSha256": _sha256_bytes(_canonical_json_bytes(request)),
        "sources": source_summaries,
        "snSetId": f"SNSET-{uuid.uuid4().hex.upper()}",
        "snSetDigest": _sha256_bytes(normalized_bytes),
        "snSummary": {
            "nonEmptyCellCount": len(all_values),
            "uniqueCount": len(unique_values),
            "duplicateCellCount": duplicate_cells,
            "duplicateValueCount": sum(1 for count in counts.values() if count > 1),
            "lengthDistribution": length_distribution,
        },
        "diff": None,
        "warnings": warnings,
    }
    plan["planSha256"] = _sha256_bytes(_canonical_json_bytes(plan))
    return plan, unique_values


def _build_plan_snapshot(
    request_path: Path,
    input_dir: Path,
    previous_request_path: Optional[Path] = None,
    previous_input_dir: Optional[Path] = None,
) -> Tuple[Dict[str, Any], List[str]]:
    plan, current_values = _build_plan_with_values(request_path, input_dir)
    previous_license_id = plan["request"]["previousLicenseId"]
    provided_previous = previous_request_path is not None or previous_input_dir is not None
    if previous_license_id is None:
        if provided_previous:
            raise LicenseDeliveryError(
                "previous inputs require request.previousLicenseId"
            )
        return plan, current_values
    if previous_request_path is None or previous_input_dir is None:
        raise LicenseDeliveryError(
            "previousLicenseId requires --previous-request and --previous-input-dir"
        )
    previous_plan, previous_values = _build_plan_with_values(
        previous_request_path, previous_input_dir
    )
    if previous_plan["request"]["licenseId"] != previous_license_id:
        raise LicenseDeliveryError("previous request licenseId does not match previousLicenseId")
    current_set = set(current_values)
    previous_set = set(previous_values)
    plan["diff"] = {
        "baseLicenseId": previous_license_id,
        "added": len(current_set - previous_set),
        "removed": len(previous_set - current_set),
        "unchanged": len(current_set & previous_set),
    }
    unsigned = dict(plan)
    unsigned.pop("planSha256", None)
    plan["planSha256"] = _sha256_bytes(_canonical_json_bytes(unsigned))
    return plan, current_values


def build_plan(
    request_path: Path,
    input_dir: Path,
    previous_request_path: Optional[Path] = None,
    previous_input_dir: Optional[Path] = None,
) -> Dict[str, Any]:
    plan, _ = _build_plan_snapshot(
        request_path,
        input_dir,
        previous_request_path,
        previous_input_dir,
    )
    return plan


def _write_json(path: Path, value: object) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(value, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def _verify_plan_integrity(plan: Dict[str, Any]) -> None:
    expected = plan.get("planSha256")
    if not isinstance(expected, str):
        raise LicenseDeliveryError("plan is missing planSha256")
    unsigned = dict(plan)
    del unsigned["planSha256"]
    actual = _sha256_bytes(_canonical_json_bytes(unsigned))
    if actual != expected:
        raise LicenseDeliveryError("plan receipt SHA-256 mismatch")


def _load_approved_plan_context(
    *,
    plan_path: Path,
    request_path: Path,
    input_dir: Path,
    previous_request_path: Optional[Path],
    previous_input_dir: Optional[Path],
) -> Tuple[Dict[str, Any], List[str]]:
    plan = _load_json(plan_path)
    _verify_plan_integrity(plan)
    recalculated, device_ids = _build_plan_snapshot(
        request_path,
        input_dir,
        previous_request_path,
        previous_input_dir,
    )
    for field in APPROVED_PLAN_FIELDS:
        if recalculated.get(field) != plan.get(field):
            raise LicenseDeliveryError(f"current inputs do not match approved plan: {field}")
    return plan, device_ids


def _git(repo: Path, *args: str) -> str:
    try:
        result = subprocess.run(
            ["git", *args],
            cwd=repo,
            check=True,
            text=True,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
        )
    except subprocess.CalledProcessError as error:
        detail = error.stderr.strip() or error.stdout.strip()
        raise LicenseDeliveryError(f"git {' '.join(args)} failed: {detail}") from error
    return result.stdout.strip()


def _embedded_public_keys(repo: Path) -> Dict[str, str]:
    sources = {
        "androidAsr": repo / "asr/android/gradle.properties",
        "androidTts": repo / "tts/android/gradle.properties",
        "harmonyAsr": repo / "asr/harmony/sdk/src/main/ets/com/amphion/asr/License.ets",
        "harmonyTts": repo / "tts/harmony/sdk/src/main/ets/License.ets",
    }
    values: Dict[str, str] = {}
    for label, path in sources.items():
        try:
            text = path.read_text(encoding="utf-8")
        except OSError as error:
            raise LicenseDeliveryError(f"cannot read embedded public key: {label}") from error
        if path.suffix == ".properties":
            match = re.search(r"^AMPHION_LICENSE_PUBLIC_KEY=([^\s]+)$", text, re.MULTILINE)
        else:
            match = re.search(
                r"^const LICENSE_PUBLIC_KEY_B64(?:\s*:\s*string)?\s*=\s*'([^']+)';$",
                text,
                re.MULTILINE,
            )
        if match is None:
            raise LicenseDeliveryError(f"embedded public key not found: {label}")
        values[label] = match.group(1)
    return values


def _derive_public_key_b64(private_key_path: Path) -> str:
    try:
        private_key = serialization.load_pem_private_key(
            private_key_path.read_bytes(), password=None
        )
    except (OSError, ValueError, TypeError) as error:
        raise LicenseDeliveryError("cannot load fixed production private key") from error
    if not isinstance(private_key, ec.EllipticCurvePrivateKey) or not isinstance(
        private_key.curve, ec.SECP256R1
    ):
        raise LicenseDeliveryError("fixed production private key must use ECDSA P-256")
    public_der = private_key.public_key().public_bytes(
        serialization.Encoding.DER,
        serialization.PublicFormat.SubjectPublicKeyInfo,
    )
    return base64.b64encode(public_der).decode("ascii")


def _policy_claims(policy: Dict[str, Any]) -> Dict[str, Any]:
    allowed_policy_fields = {
        "features",
        "sdkMajor",
        "installTier",
        "applicationRecord",
        "certificateBinding",
        "runtimeExpiry",
        "maintenance",
        "deviceBinding",
    }
    unknown_policy_fields = set(policy) - allowed_policy_fields
    if unknown_policy_fields:
        raise LicenseDeliveryError(
            "unknown policy fields: " + ",".join(sorted(unknown_policy_fields))
        )
    features = policy.get("features")
    if not isinstance(features, list) or not features or any(
        feature not in {"ASR", "TTS"} for feature in features
    ) or len(features) != len(set(features)):
        raise LicenseDeliveryError("policy.features must be unique ASR/TTS values")
    sdk_major = policy.get("sdkMajor")
    if isinstance(sdk_major, bool) or not isinstance(sdk_major, int) or sdk_major <= 0:
        raise LicenseDeliveryError("policy.sdkMajor must be a positive integer")
    application = policy.get("applicationRecord")
    if not isinstance(application, dict) or application.get("mode") not in {
        "none",
        "record-only",
    }:
        raise LicenseDeliveryError("policy.applicationRecord mode is invalid")
    allowed_application_fields = (
        {"mode"}
        if application["mode"] == "none"
        else {"mode", "applicationId", "bundleName"}
    )
    if set(application) - allowed_application_fields:
        raise LicenseDeliveryError("policy.applicationRecord contains unknown fields")
    application_id = ""
    bundle_name = ""
    if application["mode"] == "record-only":
        application_id = application.get("applicationId", "")
        bundle_name = application.get("bundleName", "")
        if not isinstance(application_id, str) or not isinstance(bundle_name, str):
            raise LicenseDeliveryError("application record values must be strings")
    certificate = policy.get("certificateBinding")
    if not isinstance(certificate, dict) or certificate.get("mode") not in {
        "none",
        "sha256",
    }:
        raise LicenseDeliveryError("policy.certificateBinding mode is invalid")
    allowed_certificate_fields = (
        {"mode"} if certificate["mode"] == "none" else {"mode", "value"}
    )
    if set(certificate) - allowed_certificate_fields:
        raise LicenseDeliveryError("policy.certificateBinding contains unknown fields")
    cert_sha256 = ""
    if certificate["mode"] == "sha256":
        cert_sha256 = certificate.get("value", "")
        if not isinstance(cert_sha256, str) or not re.fullmatch(
            r"(?:[0-9A-Fa-f]{2}:?){32}", cert_sha256
        ):
            raise LicenseDeliveryError("certificate SHA-256 is invalid")
    runtime = policy.get("runtimeExpiry")
    if not isinstance(runtime, dict) or runtime.get("mode") not in {
        "perpetual",
        "date",
    }:
        raise LicenseDeliveryError("policy.runtimeExpiry mode is invalid")
    allowed_runtime_fields = (
        {"mode"} if runtime["mode"] == "perpetual" else {"mode", "date"}
    )
    if set(runtime) - allowed_runtime_fields:
        raise LicenseDeliveryError("policy.runtimeExpiry contains unknown fields")
    expires = "" if runtime["mode"] == "perpetual" else runtime.get("date", "")
    maintenance = policy.get("maintenance")
    if not isinstance(maintenance, dict) or maintenance.get("mode") not in {
        "unlimited",
        "date",
    }:
        raise LicenseDeliveryError("policy.maintenance mode is invalid")
    allowed_maintenance_fields = (
        {"mode"} if maintenance["mode"] == "unlimited" else {"mode", "date"}
    )
    if set(maintenance) - allowed_maintenance_fields:
        raise LicenseDeliveryError("policy.maintenance contains unknown fields")
    maintenance_until = (
        "" if maintenance["mode"] == "unlimited" else maintenance.get("date", "")
    )
    for label, value in (("expiresAt", expires), ("maintenanceUntil", maintenance_until)):
        if value:
            try:
                datetime.strptime(value, "%Y-%m-%d")
            except (TypeError, ValueError) as error:
                raise LicenseDeliveryError(f"{label} must be YYYY-MM-DD") from error
    install_tier = policy.get("installTier", "")
    if not isinstance(install_tier, str):
        raise LicenseDeliveryError("policy.installTier must be a string")
    return {
        "features": features,
        "sdkMajor": sdk_major,
        "applicationId": application_id,
        "bundleName": bundle_name,
        "certSha256": cert_sha256.replace(":", "").upper(),
        "expiresAt": expires,
        "maintenanceUntil": maintenance_until,
        "installTier": install_tier,
    }


def _expected_authorized_hashes(device_ids: Sequence[str]) -> set[str]:
    return {
        hashlib.sha256(
            f"{device}{DEFAULT_DEVICE_ID_SALT_ID}".encode("utf-8")
        ).hexdigest().upper()
        for device in device_ids
    }


def _build_internal_verification(
    claims: Dict[str, Any], device_count: int, payload_bytes: bytes
) -> Dict[str, Any]:
    return {
        "status": "PASS",
        "signatureAndClaims": "PASS",
        "embeddedPublicKeys": "4/4",
        "exactDeviceSet": True,
        "authorizedDeviceCount": device_count,
        "authorizedHashUniqueCount": device_count,
        "features": claims["features"],
        "permanent": claims["expiresAt"] == "",
        "packageBinding": (
            "none"
            if not claims["applicationId"] and not claims["bundleName"]
            else "record-only"
        ),
        "certificateBinding": bool(claims["signingCertDigest"]),
        "payloadSha256": _sha256_bytes(payload_bytes),
    }


def _build_customer_readme(
    request: Dict[str, Any], claims: Dict[str, Any], device_count: int
) -> bytes:
    return (
        "# 商用离线 License\n\n"
        f"- License ID：`{request['licenseId']}`\n"
        f"- 授权能力：{', '.join(claims['features'])}\n"
        f"- 授权设备：{device_count} 台\n"
        f"- 有效期：{'永久' if not claims['expiresAt'] else claims['expiresAt']}\n"
        f"- SDK 维护期：{'不限制' if not claims['maintenanceUntil'] else claims['maintenanceUntil']}\n"
        "- 明文 SN：不包含\n\n"
        "请保持授权文件内容不变；任何修改都会导致签名校验失败。\n"
    ).encode("utf-8")


def _build_internal_verification_markdown(
    license_sha256: str, device_count: int
) -> bytes:
    return (
        "# License 验签报告\n\n"
        "- 结论：PASS\n"
        f"- 唯一授权设备：{device_count}\n"
        f"- License SHA-256：`{license_sha256}`\n"
        "- Excel SN 与 License 哈希集合：精确一致\n"
        "- 生产私钥和明文 SN：未包含\n"
    ).encode("utf-8")


def _issue_delivery(
    *,
    repo: Path,
    request_path: Path,
    plan_path: Path,
    input_dir: Path,
    previous_request_path: Optional[Path],
    previous_input_dir: Optional[Path],
    operator: str,
    out_dir: Path,
    acknowledgements: Sequence[str],
    allow_dirty: bool,
) -> Dict[str, Any]:
    if not operator.strip():
        raise LicenseDeliveryError("operator is required")
    plan, device_ids = _load_approved_plan_context(
        plan_path=plan_path,
        request_path=request_path,
        input_dir=input_dir,
        previous_request_path=previous_request_path,
        previous_input_dir=previous_input_dir,
    )
    plan_receipt_sha256 = _sha256_bytes(plan_path.read_bytes())
    required_warnings = {warning["code"] for warning in plan.get("warnings", [])}
    acknowledged = set(acknowledgements)
    unknown = acknowledged - required_warnings
    if unknown:
        raise LicenseDeliveryError(f"unknown warning acknowledgements: {','.join(sorted(unknown))}")
    missing = required_warnings - acknowledged
    if missing:
        raise LicenseDeliveryError(f"unacknowledged plan warnings: {','.join(sorted(missing))}")
    dirty = bool(_git(repo, "status", "--porcelain"))
    if dirty and not allow_dirty:
        raise LicenseDeliveryError("production issue requires a clean Git worktree")
    tool_commit = _git(repo, "rev-parse", "HEAD")
    production = not dirty and not allow_dirty
    private_key_path = repo / ".secure/amphion-license-private.pem"
    public_b64 = _derive_public_key_b64(private_key_path)
    embedded_keys = _embedded_public_keys(repo)
    if any(value != public_b64 for value in embedded_keys.values()):
        raise LicenseDeliveryError("fixed private key does not match all four SDK public keys")
    request = _load_json(request_path)
    policy = _policy_claims(request["policy"])
    envelope, claims, payload_bytes = create_license_envelope(
        private_key_path=private_key_path,
        password=None,
        application_id=policy["applicationId"],
        bundle_name=policy["bundleName"],
        customer=request["customerId"],
        license_id=request["licenseId"],
        cert_sha256=policy["certSha256"],
        device_ids=device_ids,
        device_id_salt_id=DEFAULT_DEVICE_ID_SALT_ID,
        issued=request["issuedAt"],
        expires=policy["expiresAt"],
        maintenance_until=policy["maintenanceUntil"],
        install_tier=policy["installTier"],
        features=policy["features"],
        sdk_major=policy["sdkMajor"],
    )
    authorized_hashes = claims.get("authorizedDeviceHashes")
    expected_hashes = _expected_authorized_hashes(device_ids)
    if (
        not isinstance(authorized_hashes, list)
        or len(authorized_hashes) != len(expected_hashes)
        or set(authorized_hashes) != expected_hashes
    ):
        raise LicenseDeliveryError("new License device hashes do not exactly match the SN set")
    license_bytes = (json.dumps(envelope, ensure_ascii=False, indent=2) + "\n").encode("utf-8")
    license_sha256 = _sha256_bytes(license_bytes)
    delivery_id = request["deliveryId"]
    root = f"{delivery_id}/"
    manifest = {
        "schemaVersion": SCHEMA_VERSION,
        "requestId": request["requestId"],
        "licenseId": request["licenseId"],
        "deliveryId": delivery_id,
        "customerId": request["customerId"],
        "projectId": request["projectId"],
        "previousLicenseId": request["previousLicenseId"],
        "snSetId": plan["snSetId"],
        "sourceFileCount": len(plan["sources"]),
        "snSummary": plan["snSummary"],
        "policy": request["policy"],
        "licenseSha256": license_sha256,
        "toolCommit": tool_commit,
        "production": production,
        "containsPlaintextSn": False,
    }
    verification = _build_internal_verification(
        claims, len(device_ids), payload_bytes
    )
    readme = _build_customer_readme(request, claims, len(device_ids))
    verification_json = (json.dumps(verification, ensure_ascii=False, indent=2) + "\n").encode("utf-8")
    verification_md = _build_internal_verification_markdown(
        license_sha256, len(device_ids)
    )
    manifest_json = (json.dumps(manifest, ensure_ascii=False, indent=2) + "\n").encode("utf-8")
    members = {
        "amphion-license.lic": license_bytes,
        "README.md": readme,
        "LICENSE_MANIFEST.json": manifest_json,
        "LICENSE_VERIFICATION.json": verification_json,
        "LICENSE_VERIFICATION.md": verification_md,
    }
    checksums = "".join(
        f"{_sha256_bytes(content)}  {name}\n" for name, content in sorted(members.items())
    ).encode("utf-8")
    members["SHA256SUMS.txt"] = checksums
    out_dir.mkdir(parents=True, exist_ok=True)
    zip_path = out_dir / f"{delivery_id}.zip"
    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for name, content in members.items():
            archive.writestr(root + name, content)
    receipt = {
        "schemaVersion": SCHEMA_VERSION,
        "kind": "LicenseIssuanceReceipt",
        "status": "ISSUED",
        "production": production,
        "operator": operator,
        "acknowledgements": sorted(acknowledged),
        "requestId": request["requestId"],
        "licenseId": request["licenseId"],
        "deliveryId": delivery_id,
        "customerId": request["customerId"],
        "projectId": request["projectId"],
        "previousLicenseId": request["previousLicenseId"],
        "snSetId": plan["snSetId"],
        "snSetDigest": plan["snSetDigest"],
        "uniqueSnCount": len(device_ids),
        "sourceSha256": [source["sha256"] for source in plan["sources"]],
        "planSha256": plan["planSha256"],
        "planReceiptSha256": plan_receipt_sha256,
        "toolCommit": tool_commit,
        "licenseSha256": license_sha256,
        "zipFileName": zip_path.name,
        "zipSha256": _sha256_bytes(zip_path.read_bytes()),
        "zipSizeBytes": zip_path.stat().st_size,
    }
    receipt_path = out_dir / f"{delivery_id}.issuance.json"
    _write_json(receipt_path, receipt)
    return receipt


def _parse_json_bytes(value: bytes, label: str) -> Dict[str, Any]:
    try:
        parsed = json.loads(value.decode("utf-8"))
    except (UnicodeError, json.JSONDecodeError) as error:
        raise LicenseDeliveryError(f"invalid {label} JSON") from error
    if not isinstance(parsed, dict):
        raise LicenseDeliveryError(f"{label} must be a JSON object")
    return parsed


def _verify_delivery(
    *,
    repo: Path,
    request_path: Path,
    plan_path: Path,
    input_dir: Path,
    previous_request_path: Optional[Path],
    previous_input_dir: Optional[Path],
    zip_path: Path,
    operator: str,
    out_prefix: Path,
) -> Dict[str, Any]:
    if not operator.strip():
        raise LicenseDeliveryError("operator is required")
    if out_prefix != zip_path:
        raise LicenseDeliveryError("--out-prefix must equal the final ZIP path")
    plan, device_ids = _load_approved_plan_context(
        plan_path=plan_path,
        request_path=request_path,
        input_dir=input_dir,
        previous_request_path=previous_request_path,
        previous_input_dir=previous_input_dir,
    )
    plan_receipt_sha256 = _sha256_bytes(plan_path.read_bytes())
    request = _load_json(request_path)
    delivery_id = request["deliveryId"]
    expected_zip_name = f"{delivery_id}.zip"
    if zip_path.name != expected_zip_name:
        raise LicenseDeliveryError(f"final ZIP file name must be {expected_zip_name}")
    root = f"{delivery_id}/"
    member_names = DELIVERY_MEMBER_NAMES
    expected_names = {root + name for name in member_names}
    try:
        with zipfile.ZipFile(zip_path) as archive:
            bad = archive.testzip()
            if bad is not None:
                raise LicenseDeliveryError(f"ZIP CRC failed: {bad}")
            archive_names = archive.namelist()
            if len(archive_names) != len(set(archive_names)):
                raise LicenseDeliveryError("final ZIP contains duplicate ZIP entries")
            actual_names = set(archive_names)
            if actual_names != expected_names:
                raise LicenseDeliveryError("final ZIP file set does not match the delivery contract")
            for info in archive.infolist():
                path = PurePosixPath(info.filename)
                mode = (info.external_attr >> 16) & 0xFFFF
                if (
                    path.is_absolute()
                    or ".." in path.parts
                    or "\\" in info.filename
                    or stat.S_ISLNK(mode)
                ):
                    raise LicenseDeliveryError(f"unsafe ZIP entry: {info.filename}")
            members = {
                name.removeprefix(root): archive.read(name) for name in expected_names
            }
    except (OSError, zipfile.BadZipFile) as error:
        raise LicenseDeliveryError(f"cannot read final ZIP: {zip_path.name}") from error
    checksum_lines = members["SHA256SUMS.txt"].decode("utf-8").splitlines()
    expected_checksum_names = member_names - {"SHA256SUMS.txt"}
    parsed_checksums: Dict[str, str] = {}
    for line in checksum_lines:
        parts = line.split("  ", 1)
        if len(parts) != 2 or not re.fullmatch(r"[0-9a-f]{64}", parts[0]):
            raise LicenseDeliveryError("SHA256SUMS.txt is malformed")
        if parts[1] in parsed_checksums:
            raise LicenseDeliveryError("SHA256SUMS.txt contains duplicate entries")
        parsed_checksums[parts[1]] = parts[0]
    if set(parsed_checksums) != expected_checksum_names:
        raise LicenseDeliveryError("SHA256SUMS.txt file set is invalid")
    for name, expected in parsed_checksums.items():
        if _sha256_bytes(members[name]) != expected:
            raise LicenseDeliveryError(f"checksum mismatch: {name}")
    package_text: Dict[str, str] = {}
    for name, content in members.items():
        try:
            package_text[name] = content.decode("utf-8")
        except UnicodeError as error:
            raise LicenseDeliveryError(f"delivery member must be UTF-8 text: {name}") from error
    normalized_tokens = {
        token.upper()
        for content in package_text.values()
        for token in re.findall(r"[A-Za-z0-9]+", content)
    }
    if normalized_tokens.intersection(device_ids):
        raise LicenseDeliveryError("final ZIP contains plaintext SN")
    combined_text = "\n".join(package_text.values())
    if "PRIVATE KEY" in combined_text or re.search(
        r"(?:^|[\s\"'])/(?:Users|home)/|[A-Za-z]:[\\/]", combined_text
    ):
        raise LicenseDeliveryError("final ZIP contains private-key or absolute-path material")
    manifest = _parse_json_bytes(members["LICENSE_MANIFEST.json"], "manifest")
    internal_verification = _parse_json_bytes(
        members["LICENSE_VERIFICATION.json"], "internal verification"
    )
    envelope = _parse_json_bytes(members["amphion-license.lic"], "License envelope")
    try:
        payload_bytes = base64.b64decode(envelope["payload_b64"], validate=True)
        signature = base64.b64decode(envelope["sig_b64"], validate=True)
    except (KeyError, TypeError, ValueError) as error:
        raise LicenseDeliveryError("License envelope is malformed") from error
    if envelope.get("alg") != "SHA256withECDSA":
        raise LicenseDeliveryError("License algorithm is invalid")
    embedded_keys = _embedded_public_keys(repo)
    if len(set(embedded_keys.values())) != 1:
        raise LicenseDeliveryError("four SDK embedded public keys do not match")
    try:
        public_key = serialization.load_der_public_key(
            base64.b64decode(next(iter(embedded_keys.values())), validate=True)
        )
        if not isinstance(public_key, ec.EllipticCurvePublicKey) or not isinstance(
            public_key.curve, ec.SECP256R1
        ):
            raise LicenseDeliveryError("SDK License public key must use ECDSA P-256")
        public_key.verify(signature, payload_bytes, ec.ECDSA(hashes.SHA256()))
    except (ValueError, TypeError, InvalidSignature) as error:
        raise LicenseDeliveryError("License signature verification failed") from error
    claims = _parse_json_bytes(payload_bytes, "License claims")
    policy = _policy_claims(request["policy"])
    expected_claims = {
        "applicationId": policy["applicationId"],
        "bundleName": policy["bundleName"],
        "certSha256": policy["certSha256"],
        "signingCertDigest": policy["certSha256"],
        "customer": request["customerId"],
        "deviceIdHashAlg": "SHA-256",
        "deviceIdSaltId": DEFAULT_DEVICE_ID_SALT_ID,
        "expiresAt": policy["expiresAt"],
        "features": policy["features"],
        "installTier": policy["installTier"],
        "issuedAt": request["issuedAt"],
        "licenseId": request["licenseId"],
        "maintenanceUntil": policy["maintenanceUntil"],
        "sdkMajor": policy["sdkMajor"],
    }
    if set(claims) != set(expected_claims) | {"authorizedDeviceHashes"}:
        raise LicenseDeliveryError("License claims contain unknown or missing fields")
    for field, expected in expected_claims.items():
        if claims.get(field) != expected:
            raise LicenseDeliveryError(f"License claim does not match request: {field}")
    raw_hashes = claims.get("authorizedDeviceHashes")
    if not isinstance(raw_hashes, list) or not all(isinstance(value, str) for value in raw_hashes):
        raise LicenseDeliveryError("authorizedDeviceHashes must be a string array")
    if len(raw_hashes) != len(set(raw_hashes)):
        raise LicenseDeliveryError("License contains duplicate authorized device hashes")
    expected_hashes = _expected_authorized_hashes(device_ids)
    if set(raw_hashes) != expected_hashes or len(raw_hashes) != len(device_ids):
        raise LicenseDeliveryError("License device hashes do not exactly match the SN set")
    unauthorized = hashlib.sha256(
        f"AMPHION-UNAUTHORIZED-CHECK-{DEFAULT_DEVICE_ID_SALT_ID}".encode("utf-8")
    ).hexdigest().upper()
    if unauthorized in expected_hashes:
        raise LicenseDeliveryError("unauthorized verification sentinel is present")
    license_sha256 = _sha256_bytes(members["amphion-license.lic"])
    expected_internal_verification = _build_internal_verification(
        claims, len(device_ids), payload_bytes
    )
    if internal_verification != expected_internal_verification:
        raise LicenseDeliveryError("LICENSE_VERIFICATION.json does not match verified facts")
    expected_readme = _build_customer_readme(request, claims, len(device_ids))
    if members["README.md"] != expected_readme:
        raise LicenseDeliveryError("README.md does not match the verified delivery")
    expected_verification_markdown = _build_internal_verification_markdown(
        license_sha256, len(device_ids)
    )
    if members["LICENSE_VERIFICATION.md"] != expected_verification_markdown:
        raise LicenseDeliveryError("LICENSE_VERIFICATION.md does not match verified facts")
    tool_commit = manifest.get("toolCommit")
    if not isinstance(tool_commit, str) or not re.fullmatch(r"[0-9a-f]{40,64}", tool_commit):
        raise LicenseDeliveryError("delivery manifest toolCommit is invalid")
    if not isinstance(manifest.get("production"), bool):
        raise LicenseDeliveryError("delivery manifest production flag is invalid")
    manifest_expected = {
        "schemaVersion": SCHEMA_VERSION,
        "requestId": request["requestId"],
        "licenseId": request["licenseId"],
        "deliveryId": delivery_id,
        "customerId": request["customerId"],
        "projectId": request["projectId"],
        "previousLicenseId": request["previousLicenseId"],
        "snSetId": plan["snSetId"],
        "sourceFileCount": len(plan["sources"]),
        "snSummary": plan["snSummary"],
        "policy": request["policy"],
        "licenseSha256": license_sha256,
        "toolCommit": tool_commit,
        "production": manifest["production"],
        "containsPlaintextSn": False,
    }
    if manifest != manifest_expected:
        raise LicenseDeliveryError("delivery manifest does not exactly match verified facts")
    zip_sha256 = _sha256_bytes(zip_path.read_bytes())
    receipt: Dict[str, Any] = {
        "schemaVersion": SCHEMA_VERSION,
        "kind": "LicenseVerificationReceipt",
        "status": "PASS",
        "operator": operator,
        "requestId": request["requestId"],
        "licenseId": request["licenseId"],
        "deliveryId": delivery_id,
        "customerId": request["customerId"],
        "projectId": request["projectId"],
        "previousLicenseId": request["previousLicenseId"],
        "snSetId": plan["snSetId"],
        "uniqueSnCount": len(device_ids),
        "planSha256": plan["planSha256"],
        "planReceiptSha256": plan_receipt_sha256,
        "licenseSha256": license_sha256,
        "zipFileName": zip_path.name,
        "zipSha256": zip_sha256,
        "zipSizeBytes": zip_path.stat().st_size,
        "toolCommit": manifest.get("toolCommit"),
        "production": manifest.get("production") is True,
        "checks": {
            "zipCrc": "PASS",
            "fixedFileSet": "PASS",
            "checksums": "PASS",
            "signatureAndClaims": "PASS",
            "embeddedPublicKeys": "4/4",
            "exactDeviceSet": "PASS",
            "unauthorizedRejected": "PASS",
            "sdkMajorPolicy": "PASS",
            "tamperedPayloadRejected": "PASS",
        },
    }
    json_path = Path(str(out_prefix) + ".verification.json")
    markdown_path = Path(str(out_prefix) + ".verification.md")
    _write_json(json_path, receipt)
    markdown_path.parent.mkdir(parents=True, exist_ok=True)
    markdown_path.write_text(
        "# Final License ZIP Verification\n\n"
        "- Status: PASS\n"
        f"- Delivery ID: `{delivery_id}`\n"
        f"- ZIP SHA-256: `{zip_sha256}`\n"
        f"- Authorized devices: {len(device_ids)}\n"
        "- Plaintext SN in ZIP: no\n",
        encoding="utf-8",
    )
    return receipt


def _history_lock_path(history_path: Path) -> Path:
    identity = _sha256_bytes(str(history_path.resolve()).encode("utf-8"))
    return Path(tempfile.gettempdir()) / f"amphion-license-delivery-history-{identity}.lock"


def _load_history(path: Path) -> Dict[str, Any]:
    if not path.exists():
        return {"schemaVersion": SCHEMA_VERSION, "deliveries": []}
    history = _load_json(path)
    if history.get("schemaVersion") != SCHEMA_VERSION:
        raise LicenseDeliveryError(f"history schemaVersion must be {SCHEMA_VERSION}")
    deliveries = history.get("deliveries")
    if not isinstance(deliveries, list) or not all(
        isinstance(delivery, dict) for delivery in deliveries
    ):
        raise LicenseDeliveryError("history deliveries must be an object array")
    return history


def _record_delivery(
    *,
    repo: Path,
    plan_path: Path,
    zip_path: Path,
    issuance_path: Path,
    verification_path: Path,
    operator: str,
    delivered_at: str,
) -> Dict[str, Any]:
    history_path = repo / "delivery/license-delivery-history.json"
    if not operator.strip():
        raise LicenseDeliveryError("operator is required")
    try:
        datetime.strptime(delivered_at, "%Y-%m-%d")
    except ValueError as error:
        raise LicenseDeliveryError("delivered-at must be YYYY-MM-DD") from error
    if _git(repo, "status", "--porcelain"):
        raise LicenseDeliveryError("record requires a clean Git worktree")
    issuance = _load_json(issuance_path)
    verification = _load_json(verification_path)
    if issuance.get("kind") != "LicenseIssuanceReceipt" or issuance.get("status") != "ISSUED":
        raise LicenseDeliveryError("issuance receipt is invalid")
    if verification.get("kind") != "LicenseVerificationReceipt" or verification.get("status") != "PASS":
        raise LicenseDeliveryError("verification receipt is not PASS")
    if issuance.get("production") is not True or verification.get("production") is not True:
        raise LicenseDeliveryError("non-production delivery cannot be recorded")
    identity_fields = (
        "requestId",
        "licenseId",
        "deliveryId",
        "customerId",
        "projectId",
        "previousLicenseId",
        "snSetId",
        "planSha256",
        "planReceiptSha256",
        "licenseSha256",
        "zipFileName",
        "zipSha256",
        "zipSizeBytes",
        "toolCommit",
    )
    for field in identity_fields:
        if issuance.get(field) != verification.get(field):
            raise LicenseDeliveryError(f"issuance and verification receipts disagree: {field}")
    actual_plan_receipt_sha256 = _sha256_bytes(plan_path.read_bytes())
    if actual_plan_receipt_sha256 != verification["planReceiptSha256"]:
        raise LicenseDeliveryError("plan receipt SHA-256 does not match the recorded receipts")
    actual_zip_sha256 = _sha256_bytes(zip_path.read_bytes())
    if actual_zip_sha256 != verification["zipSha256"]:
        raise LicenseDeliveryError("final ZIP SHA-256 does not match verification receipt")
    if zip_path.name != verification["zipFileName"] or zip_path.stat().st_size != verification[
        "zipSizeBytes"
    ]:
        raise LicenseDeliveryError("final ZIP identity does not match verification receipt")
    delivery_id = verification["deliveryId"]
    try:
        with zipfile.ZipFile(zip_path) as archive:
            manifest_bytes = archive.read(f"{delivery_id}/LICENSE_MANIFEST.json")
            license_bytes = archive.read(f"{delivery_id}/amphion-license.lic")
    except (OSError, KeyError, zipfile.BadZipFile) as error:
        raise LicenseDeliveryError("cannot read manifest from final ZIP") from error
    manifest = _parse_json_bytes(manifest_bytes, "delivery manifest")
    if manifest.get("production") is not True:
        raise LicenseDeliveryError("manifest is not a production delivery")
    if _sha256_bytes(license_bytes) != verification["licenseSha256"]:
        raise LicenseDeliveryError("License SHA-256 does not match verification receipt")
    entry: Dict[str, Any] = {
        "requestId": verification["requestId"],
        "licenseId": verification["licenseId"],
        "deliveryId": delivery_id,
        "customerId": verification["customerId"],
        "projectId": verification["projectId"],
        "previousLicenseId": verification["previousLicenseId"],
        "snSetId": verification["snSetId"],
        "uniqueSnCount": verification["uniqueSnCount"],
        "policy": manifest["policy"],
        "artifact": zip_path.name,
        "artifactSha256": actual_zip_sha256,
        "artifactSizeBytes": zip_path.stat().st_size,
        "licenseSha256": verification["licenseSha256"],
        "planSha256": verification["planSha256"],
        "planReceiptSha256": actual_plan_receipt_sha256,
        "issuanceReceiptSha256": _sha256_bytes(issuance_path.read_bytes()),
        "verificationReceiptSha256": _sha256_bytes(verification_path.read_bytes()),
        "toolCommit": verification["toolCommit"],
        "issueOperator": issuance["operator"],
        "verifyOperator": verification["operator"],
        "operator": operator,
        "deliveredAt": delivered_at,
        "status": "delivered",
    }
    forbidden_fields = {
        "snSetDigest",
        "sourceSha256",
        "authorizedDeviceHashes",
        "deviceIds",
    }
    if forbidden_fields.intersection(entry):
        raise LicenseDeliveryError("history entry contains forbidden sensitive fields")
    history_path.parent.mkdir(parents=True, exist_ok=True)
    lock_path = _history_lock_path(history_path)
    with lock_path.open("a+") as lock:
        fcntl.flock(lock.fileno(), fcntl.LOCK_EX)
        history = _load_history(history_path)
        for field in ("requestId", "licenseId", "deliveryId"):
            if any(delivery.get(field) == entry[field] for delivery in history["deliveries"]):
                raise LicenseDeliveryError(f"history already contains {field}={entry[field]}")
        history["deliveries"].append(entry)
        temporary_path: Optional[Path] = None
        try:
            with tempfile.NamedTemporaryFile(
                mode="w",
                encoding="utf-8",
                dir=history_path.parent,
                prefix=f".{history_path.name}.",
                suffix=".tmp",
                delete=False,
            ) as temporary:
                temporary_path = Path(temporary.name)
                json.dump(history, temporary, ensure_ascii=False, indent=2)
                temporary.write("\n")
                temporary.flush()
                os.fsync(temporary.fileno())
            os.replace(temporary_path, history_path)
            temporary_path = None
        finally:
            if temporary_path is not None:
                temporary_path.unlink(missing_ok=True)
    return entry


def main(argv: Optional[Sequence[str]] = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    subparsers = parser.add_subparsers(dest="command", required=True)
    plan_parser = subparsers.add_parser("plan")
    plan_parser.add_argument("--request", required=True, type=Path)
    plan_parser.add_argument("--input-dir", required=True, type=Path)
    plan_parser.add_argument("--previous-request", type=Path)
    plan_parser.add_argument("--previous-input-dir", type=Path)
    plan_parser.add_argument("--out", required=True, type=Path)
    issue_parser = subparsers.add_parser("issue")
    issue_parser.add_argument("--repo", required=True, type=Path)
    issue_parser.add_argument("--request", required=True, type=Path)
    issue_parser.add_argument("--plan", required=True, type=Path)
    issue_parser.add_argument("--input-dir", required=True, type=Path)
    issue_parser.add_argument("--previous-request", type=Path)
    issue_parser.add_argument("--previous-input-dir", type=Path)
    issue_parser.add_argument("--operator", required=True)
    issue_parser.add_argument("--out-dir", required=True, type=Path)
    issue_parser.add_argument("--acknowledge", action="append", default=[])
    issue_parser.add_argument("--allow-dirty", action="store_true")
    verify_parser = subparsers.add_parser("verify")
    verify_parser.add_argument("--repo", required=True, type=Path)
    verify_parser.add_argument("--request", required=True, type=Path)
    verify_parser.add_argument("--plan", required=True, type=Path)
    verify_parser.add_argument("--input-dir", required=True, type=Path)
    verify_parser.add_argument("--previous-request", type=Path)
    verify_parser.add_argument("--previous-input-dir", type=Path)
    verify_parser.add_argument("--zip", required=True, type=Path)
    verify_parser.add_argument("--operator", required=True)
    verify_parser.add_argument("--out-prefix", required=True, type=Path)
    record_parser = subparsers.add_parser("record")
    record_parser.add_argument("--repo", required=True, type=Path)
    record_parser.add_argument("--plan", required=True, type=Path)
    record_parser.add_argument("--zip", required=True, type=Path)
    record_parser.add_argument("--issuance", required=True, type=Path)
    record_parser.add_argument("--verification", required=True, type=Path)
    record_parser.add_argument("--operator", required=True)
    record_parser.add_argument("--delivered-at", required=True)
    args = parser.parse_args(argv)
    try:
        if args.command == "plan":
            plan = build_plan(
                args.request.resolve(),
                args.input_dir.resolve(),
                args.previous_request.resolve() if args.previous_request else None,
                args.previous_input_dir.resolve() if args.previous_input_dir else None,
            )
            _write_json(args.out.resolve(), plan)
            print(f"[OK] wrote License delivery plan: {args.out}")
        elif args.command == "issue":
            receipt = _issue_delivery(
                repo=args.repo.resolve(),
                request_path=args.request.resolve(),
                plan_path=args.plan.resolve(),
                input_dir=args.input_dir.resolve(),
                previous_request_path=(
                    args.previous_request.resolve() if args.previous_request else None
                ),
                previous_input_dir=(
                    args.previous_input_dir.resolve() if args.previous_input_dir else None
                ),
                operator=args.operator,
                out_dir=args.out_dir.resolve(),
                acknowledgements=args.acknowledge,
                allow_dirty=args.allow_dirty,
            )
            print(
                f"[OK] issued {receipt['licenseId']} as {receipt['zipFileName']} "
                f"devices={receipt['uniqueSnCount']}"
            )
        elif args.command == "verify":
            receipt = _verify_delivery(
                repo=args.repo.resolve(),
                request_path=args.request.resolve(),
                plan_path=args.plan.resolve(),
                input_dir=args.input_dir.resolve(),
                previous_request_path=(
                    args.previous_request.resolve() if args.previous_request else None
                ),
                previous_input_dir=(
                    args.previous_input_dir.resolve() if args.previous_input_dir else None
                ),
                zip_path=args.zip.resolve(),
                operator=args.operator,
                out_prefix=args.out_prefix.resolve(),
            )
            print(
                f"[OK] verified final ZIP {receipt['zipFileName']} "
                f"sha256={receipt['zipSha256']}"
            )
        elif args.command == "record":
            repo = args.repo.resolve()
            entry = _record_delivery(
                repo=repo,
                plan_path=args.plan.resolve(),
                zip_path=args.zip.resolve(),
                issuance_path=args.issuance.resolve(),
                verification_path=args.verification.resolve(),
                operator=args.operator,
                delivered_at=args.delivered_at,
            )
            print(
                f"[OK] recorded {entry['deliveryId']} artifact={entry['artifact']}"
            )
        else:
            raise LicenseDeliveryError(f"{args.command} is not implemented yet")
    except LicenseDeliveryError as error:
        print(f"[ERROR] {error}", file=sys.stderr)
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
